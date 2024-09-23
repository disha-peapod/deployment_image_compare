import os
import yaml
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pdfkit

# Define the base paths for dev and qa environments
base_path_dev = "pdl-coreservices-app-deployments/app-values-02/nonprd"
base_path_qa = "pdl-coreservices-app-deployments/app-values-02/nonprd"

# Function to get image tag from a YAML file
def get_image_tag(file_path):
    try:
        with open(file_path, 'r') as file:
            data = yaml.safe_load(file)
            return data.get('image', {}).get('tag', 'Not Found')
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return 'Error'
os.chdir('..') 
# Prepare the table data
table_data = []

# Get the list of service names (folder names) in the base path
service_names = [name for name in os.listdir(base_path_dev) if os.path.isdir(os.path.join(base_path_dev, name))]

for service_name in service_names:
    # Construct paths for dev and qa
    dev_path = os.path.join(base_path_dev, service_name, 'centralus', 'dev', 'mrtn2')
    qa_path = os.path.join(base_path_qa, service_name, 'centralus', 'qa', 'mrtn2')
    
    # Check if both dev and qa paths exist
    if os.path.exists(dev_path) and os.path.exists(qa_path):
        # Get all YAML files in both directories
        dev_files = {f for f in os.listdir(dev_path) if f.endswith('.yaml')}
        qa_files = {f for f in os.listdir(qa_path) if f.endswith('.yaml')}
        
        # Find matching files in both environments
        matching_files = dev_files.intersection(qa_files)
        
        for filename in matching_files:
            dev_image_tag = get_image_tag(os.path.join(dev_path, filename))
            qa_image_tag = get_image_tag(os.path.join(qa_path, filename))
            
            # Add the row to the table
            table_data.append({
                'Service Name': service_name,
                'Filename': filename,
                'Dev Image Tag': dev_image_tag,
                'QA Image Tag': qa_image_tag
            })

# Create a DataFrame
df = pd.DataFrame(table_data)
df['compare'] = df.apply(
    lambda row: 'Match' if row['Dev Image Tag'] == row['QA Image Tag'] else 
                'Dev > QA' if row['Dev Image Tag'] > row['QA Image Tag'] else 
                'Dev < QA', axis=1
)


os.chdir('deployment_image_compare')

output_filename = 'dev_qa_image_tags_compare.xlsx'
# Save DataFrame to Excel
df.to_excel(output_filename, index=False)

workbook = load_workbook(output_filename)
worksheet = workbook.active

# Define fill colors
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red

# Apply conditional formatting based on the comparison of Dev and QA image tags
for row in range(2, len(df) + 2):  # Skip header row
    dev_tag = worksheet[f'C{row}'].value  # Dev tag assumed in column C
    qa_tag = worksheet[f'D{row}'].value   # QA tag assumed in column D

    # Conditional formatting logic
    if dev_tag == qa_tag:
        worksheet[f'E{row}'].fill = green_fill  # Optional: add text in the compare column
        worksheet[f'B{row}'].fill = green_fill  # Color filename green
    elif dev_tag > qa_tag:
        worksheet[f'E{row}'].fill = yellow_fill  # Optional: add text in the compare column
        worksheet[f'B{row}'].fill = yellow_fill  # Color filename yellow
    else:
        worksheet[f'E{row}'].fill = red_fill  # Optional: add text in the compare column
        worksheet[f'B{row}'].fill = red_fill  # Color filename red

workbook.save(output_filename)

# Save CSV
csv_filename = 'dev_qa_image_tags_csv.csv'
df.to_csv(csv_filename, index=False)

print("Make sure you have the latest changes from the main branch for - pdl-coreservices-app-deployments")
print(f"Comparison table saved as {output_filename}.")
print("RED - Dev image tag >  QA image tag")
print("YELLOW - Dev image tag < QA image tag")
print("GREEN - Dev image tag =  QA image tag")
    